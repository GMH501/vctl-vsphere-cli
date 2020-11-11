import os
import sys
from datetime import timedelta

from pyVmomi import vim, vmodl
import click
import openpyxl

from vctl.helpers.helpers import load_context, get_path
from vctl.helpers.vmware import objectsInFolder, yeldVmsInFolder, yeldVmsInCluster, get_obj
from vctl.helpers.auth import inject_token
from vctl.helpers.perf import BuildQuery, Perf, get_vms
from vctl.exceptions.exceptions import ContextNotFound


@click.group()
def perfs():
    """
    """
    pass


@perfs.command()
@click.option('--context', '-c',
              help='The context to use for run the command, the default is <current-context>.',
              required=False)
@click.option('--source', '-s',
              help='Path of the source file containing the vms, one per line.',
              required=True)
@click.option('--output', '-o',
              help='Path of the destination .xlsx generated file.',
              required=True)
@click.option('--interval', '-i',
              help='The desired interval for extract the metrics.',
              type=click.Choice(['minutes', 'hours', 'weeks']),
              required=True)
@click.option('--unit', '-u',
              help='Unit of the interval (ex. interval = weeks, unit = 4, is equal to 28 days).',
              required=True)
@click.option('--metrics', '-m', multiple=True)
def export(context, source, output, metrics, interval, unit):
    try:
        context = load_context(context=context)
        si = inject_token(context)
        content = si.content
        filename = get_path(source)
        with open(filename, "r") as f:
            vms = f.read().strip().split()
            total = len(vms)
        vc_vms = get_vms(content)
        working_path = os.getcwd()
        xls_path = get_path(output)
        wb = openpyxl.Workbook()
        iteration = 0
        alerts = ""
        for vm in vc_vms:
            if len(vms) == 0:
                wb.save(xls_path)
                sys.exit(0)
            if vm.name in vms:   
                perfResults = BuildQuery(
                                    content,
                                    metrics,
                                    vm,
                                    time_measure=interval,
                                    unit=int(unit))
                if not perfResults:
                    alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": is powered-off.".format(vm.name)
                    vms.remove(vm.name)
                    iteration += 1
                    percents = "{0:.1f}".format(100 * (iteration / total))
                    sys.stdout.write('\r{}%'.format(percents))
                    continue
                else:
                    results = Perf(content, perfResults)
                    num_metrics = results.num_metrics
                    ws = wb.create_sheet(vm.name)
                    intestazione = [(counter["counterName"] + " in " + counter["counterUnit"] + " for vm " + vm.name) for counter in results.values]
                    ws.append(["TIME"] + intestazione)
                    for index in range(0, results.num_timestamps):
                        time = results.timestamps[index] + timedelta(hours=1)
                        time = time.strftime("%a %b %d %H:%M")
                        valori = [results.values[metric]["values"][index] for metric in range(0, num_metrics)]
                        ws.append([time] + valori)
                vms.remove(vm.name)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
                wb.save(xls_path)
            else:
                continue
        if vms != []:
            for vm in vms:
                alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": not found.".format(vm)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
        print(alerts)
        wb.save(xls_path)

    except ContextNotFound:
        print('Context not found.')
        raise SystemExit()
    except vim.fault.NotAuthenticated:
        print('Context expired.')
        raise SystemExit(1)
    except vmodl.MethodFault as e:
        print('Caught vmodl fault: {}'.format(e.msg))
        raise SystemExit(1)
    except Exception as e:
        print('Caught error: {}'.format(e))
        raise SystemExit(1)


def add_column(ws, column):
    new_column = ws.max_column + 1
    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)


@perfs.command()
@click.option('--context', '-c',
              help='The context to use for run the command, the default is <current-context>.',
              required=False)
@click.option('--source', '-s',
              help='Path of the source file containing the vms, one per line.',
              required=False)
@click.option('--output', '-o',
              help='Path of the destination .xlsx generated file.',
              required=True)
@click.option('--interval', '-i',
              help='The desired interval for extract the metrics.',
              type=click.Choice(['minutes', 'hours', 'weeks']),
              required=True)
@click.option('--unit', '-u',
              help='Unit of the interval (ex. interval = weeks, unit = 4, is equal to 28 days).',
              required=True)
@click.option('--metrics', '-m', multiple=True)
def exports(context, source, output, metrics, interval, unit):
    try:
        context = load_context(context=context)
        si = inject_token(context)
        content = si.content
        filename = get_path(source)
        with open(filename, "r") as f:
            vms = f.read().strip().split()
            total = len(vms)
        vc_vms = get_vms(content)
        xls_path = get_path(output)
        wb = openpyxl.Workbook()
        for metric in metrics:
            wb.create_sheet(metric)
        iteration = 0
        alerts = ""
        firstVm = True
        for vm in vc_vms:
            if len(vms) == 0:
                wb.save(xls_path)
                sys.exit(0)
            if vm.name in vms:       
                perfResults = BuildQuery(
                                    content,
                                    metrics, 
                                    vm, 
                                    time_measure=interval, 
                                    unit=int(unit))
                if not perfResults or perfResults[0].sampleInfo == []:
                    alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": is powered-off.".format(vm.name)
                    vms.remove(vm.name)
                    iteration += 1
                    percents = "{0:.1f}".format(100 * (iteration / total))
                    sys.stdout.write('\r{}%'.format(percents))
                    continue
                else:
                    results = Perf(content, perfResults)
                    num_metrics = results.num_metrics
                    for counter in results.values:
                        column = []
                        ws = wb[counter["counterName"]]
                        intestazione = [(vm.name)]
                        if firstVm:
                            ws.append(["TIME"] + intestazione)
                            for index in range (0, results.num_timestamps):
                                time = results.timestamps[index] + timedelta(hours=1)
                                time = time.strftime("%a %b %d %H:%M")
                                valore = counter["values"][index]
                                ws.append([time] + [valore])
                        else:
                            column = [vm.name]
                            column += counter["values"]
                            add_column(ws, column)
        
                vms.remove(vm.name)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
                wb.save(xls_path)
                firstVm = False
            else:
                continue
        if vms != []:
            for vm in vms:
                alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": not found.".format(vm)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
        print(alerts)
        wb.save(xls_path)

    except ContextNotFound:
        print('Context not found.')
        raise SystemExit()
    except vim.fault.NotAuthenticated:
        print('Context expired.')
        raise SystemExit(1)
    except vmodl.MethodFault as e:
        print('Caught vmodl fault: {}'.format(e.msg))
        raise SystemExit(1)
    except Exception as e:
        print('Caught error: {}'.format(e))
        raise SystemExit(1)



@perfs.command()
@click.option('--context', '-c',
              help='The context to use for run the command, the default is <current-context>.',
              required=False)
@click.option('--folder', '-f',
              help='Path of the source file containing the vms, one per line.',
              required=True)
@click.option('--output', '-o',
              help='Path of the destination .xlsx generated file.',
              required=True)
@click.option('--interval', '-i',
              help='The desired interval for extract the metrics.',
              type=click.Choice(['minutes', 'hours', 'weeks']),
              required=True)
@click.option('--unit', '-u',
              help='Unit of the interval (ex. interval = weeks, unit = 4, is equal to 28 days).',
              required=True)
@click.option('--metrics', '-m', multiple=True)
def fromfolder(context, folder, output, metrics, interval, unit):
    try:
        context = load_context(context=context)
        si = inject_token(context)
        content = si.content
        folderObjects = objectsInFolder(content, folder)
        vms = yeldVmsInFolder(folderObjects)
        xls_path = get_path(output)
        wb = openpyxl.Workbook()
        for metric in metrics:
            wb.create_sheet(metric)
        alerts = ""
        firstVm = True
        for vm in vms:     
            perfResults = BuildQuery(
                                content,
                                metrics, 
                                vm, 
                                time_measure=interval, 
                                unit=int(unit))
            if not perfResults or perfResults[0].sampleInfo == []:
                alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": is powered-off.".format(vm.name)
                continue
            else:
                results = Perf(content, perfResults)
                num_metrics = results.num_metrics
                for counter in results.values:
                    column = []
                    ws = wb[counter["counterName"]]
                    intestazione = [(counter["counterName"] + " in " + counter["counterUnit"] + " for vm " + vm.name)]
                    if firstVm:
                        ws.append(["TIME"] + intestazione)
                        for index in range (0, results.num_timestamps):
                            time = results.timestamps[index] + timedelta(hours=1)
                            time = time.strftime("%a %b %d %H:%M")
                            valori = counter["values"][index]
                            ws.append([time] + [valori])
                    else:
                        column = [counter["counterName"] + " in " + counter["counterUnit"] + " for vm " + vm.name]
                        column += counter["values"]
                        add_column(ws, column)
            wb.save(xls_path)
            firstVm = False
        print(alerts)
        wb.save(xls_path)

    except ContextNotFound:
        print('Context not found.')
        raise SystemExit()
    except vim.fault.NotAuthenticated:
        print('Context expired.')
        raise SystemExit(1)
    except vmodl.MethodFault as e:
        print('Caught vmodl fault: {}'.format(e.msg))
        raise SystemExit(1)
    except Exception as e:
        print('Caught error: {}'.format(e))
        raise SystemExit(1)


@perfs.command()
@click.option('--context', '-c',
              help='The context to use for run the command, the default is <current-context>.',
              required=False)
@click.option('--source', '-s',
              help='Path of the source file containing the vms, one per line.',
              required=True)
@click.option('--output', '-o',
              help='Path of the destination .xlsx generated file.',
              required=True)
@click.option('--interval', '-i',
              help='The desired interval for extract the metrics.',
              type=click.Choice(['minutes', 'hours', 'weeks']),
              required=True)
@click.option('--unit', '-u',
              help='Unit of the interval (ex. interval = weeks, unit = 4, is equal to 28 days).',
              required=True)
def xxmemexport(context, source, output, interval, unit):
    try:
        context = load_context(context=context)
        si = inject_token(context)
        content = si.content
        filename = get_path(source)
        with open(filename, "r") as f:
            vms = f.read().strip().split()
            total = len(vms)
        vc_vms = get_vms(content)
        working_path = os.getcwd()
        xls_path = get_path(output)
        wb = openpyxl.Workbook()
        wb.create_sheet('mem.usage.average')
        iteration = 0
        alerts = ""
        firstVm = True
        for vm in vc_vms:
            if len(vms) == 0:
                wb.save(xls_path)
                sys.exit(0)
            if vm.name in vms:       
                perfResults = BuildQuery(
                                    content,
                                    ['mem.usage.average'], 
                                    vm, 
                                    time_measure=interval, 
                                    unit=int(unit))
                if not perfResults or perfResults[0].sampleInfo == []:
                    alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": is powered-off.".format(vm.name)
                    vms.remove(vm.name)
                    iteration += 1
                    percents = "{0:.1f}".format(100 * (iteration / total))
                    sys.stdout.write('\r{}%'.format(percents))
                    continue
                else:
                    results = Perf(content, perfResults)
                    num_metrics = results.num_metrics
                    for counter in results.values:
                        column = []
                        ws = wb['mem.usage.average']
                        intestazione = [(vm.name)]
                        if firstVm:
                            ws.append(["TIME"] + intestazione)
                            for index in range (0, results.num_timestamps):
                                time = results.timestamps[index] + timedelta(hours=1)
                                time = time.strftime("%a %b %d %H:%M")
                                percent = counter["values"][index]
                                valore = round(percent * vm.config.hardware.memoryMB / 100, 2)
                                ws.append([time] + [valore])
                        else:
                            column = [vm.name]
                            values = [round(value * vm.config.hardware.memoryMB / 100, 2) for value in counter["values"]]
                            column += values
                            add_column(ws, column)
        
                vms.remove(vm.name)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
                wb.save(xls_path)
                firstVm = False
            else:
                continue
        if vms != []:
            for vm in vms:
                alerts = "\n[ALERT] Cannot retrieve metrics from vm \"{}\": not found.".format(vm)
                iteration += 1
                percents = "{0:.1f}".format(100 * (iteration / total))
                sys.stdout.write('\r{}%'.format(percents))
        print(alerts)
        wb.save(xls_path)

    except ContextNotFound:
        print('Context not found.')
        raise SystemExit()
    except vim.fault.NotAuthenticated:
        print('Context expired.')
        raise SystemExit(1)
    except vmodl.MethodFault as e:
        print('Caught vmodl fault: {}'.format(e.msg))
        raise SystemExit(1)
    except Exception as e:
        print('Caught error: {}'.format(e))
        raise SystemExit(1)


@perfs.command()
@click.option('--context', '-c',
              help='The context to use for run the command, the default is <current-context>.',
              required=False)
@click.option('--cluster',
              help='Path of the source file containing the vms, one per line.',
              required=False)
@click.option('--output', '-o',
              help='Path of the destination .xlsx generated file.',
              required=True)
@click.option('--interval', '-i',
              help='The desired interval for extract the metrics.',
              type=click.Choice(['minutes', 'hours', 'weeks']),
              required=True)
@click.option('--unit', '-u',
              help='Unit of the interval (ex. interval = weeks, unit = 4, is equal to 28 days).',
              required=True)
@click.option('--metrics', '-m', multiple=True)
def memexport(context, cluster, output, metrics, interval, unit):
    try:
        context = load_context(context=context)
        si = inject_token(context)
        content = si.content
        cluster = get_obj(content, [vim.ComputeResource], cluster)
        vms = yeldVmsInCluster(cluster.host)
        xls_path = get_path(output)
        wb = openpyxl.Workbook()
        wb.create_sheet('mem.usage.average')
        alerts = ""
        firstVm = True
        for vm in vms:
            if vm.summary.runtime.powerState != 'poweredOff':
                perfResults = BuildQuery(
                                    content,
                                    ['mem.usage.average'],
                                    vm,
                                    time_measure=interval,
                                    unit=int(unit))
                if not perfResults or perfResults[0].sampleInfo == []:
                    print("[ALERT] Cannot retrieve metrics from vm \"{}\", status: {}.".format(vm.name, vm.summary.runtime.powerState))
                    continue
                else:
                    print("exporting vm {}".format(vm.name))
                    results = Perf(content, perfResults)
                    for counter in results.values:
                        column = []
                        ws = wb['mem.usage.average']
                        intestazione = [(vm.name)]
                        if firstVm:
                            ws.append(["TIME"] + intestazione)
                            for index in range(0, results.num_timestamps):
                                time = results.timestamps[index] + timedelta(hours=1)
                                time = time.strftime("%a %b %d %H:%M")
                                percent = counter["values"][index]
                                valore = round(percent * vm.config.hardware.memoryMB / 100, 2)
                                ws.append([time] + [valore])
                        else:
                            column = [vm.name]
                            values = [round(value * vm.config.hardware.memoryMB / 100, 2) for value in counter["values"]]
                            column += values
                            add_column(ws, column)
                wb.save(xls_path)
                firstVm = False
            else:
                print("[ALERT] Cannot retrieve metrics from vm \"{}\", status: {}.".format(vm.name, vm.summary.runtime.powerState))
        wb.save(xls_path)

    except ContextNotFound:
        print('Context not found.')
        raise SystemExit()
    except vim.fault.NotAuthenticated:
        print('Context expired.')
        raise SystemExit(1)
    except vmodl.MethodFault as e:
        print('Caught vmodl fault: {}'.format(e.msg))
        raise SystemExit(1)
    except Exception as e:
        print('Caught error: {}'.format(e))
        raise SystemExit(1)